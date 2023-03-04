from openpyxl import Workbook
from openpyxl import load_workbook
import os

excel_file = "scrap_linkedin_posts.xslx"
if os.path.exists(excel_file): #eğer dosya varsa
    wb = load_workbook(excel_file)
    ws = wb.active

else:
    wb = Workbook()
    ws = wb.active
    ws.append([
        "Owner name",
        "Owner Url",
        "Date",
        "Text",
        "Shared text",

    ])

wb.save(excel_file)